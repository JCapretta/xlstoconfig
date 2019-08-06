#!/usr/bin/python

"""
Module docstring
"""

import argparse
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from jinja2 import Environment, FileSystemLoader
from jinja2 import contextfilter
import inspect
import custom_tests
import custom_filters
import warnings


class MyParser(argparse.ArgumentParser):
    """This class inherits from ArgumentParser and redefines error()"""
    def error(self, message):
        """override error method to print help and error message"""
        sys.stderr.write('error: %s\n' % message)
        self.print_help()
        sys.exit(2)

def get_args():
    """parse arguments from command line"""
    parser = MyParser(description="John's config generator")
    parser.add_argument('file', help='excel file')
    parser.add_argument('template', help='Jinja2 template file')
    
    if not sys.argv:
    #print help if no arguments
        parser.print_help()
        sys.exit(1)
    parsed_args = parser.parse_args()
    return parsed_args

def read_xls_dict(input_file):
    'Read the XLS file and return as a list of dictionaries:'
    spreadsheet = {}
    try: 
        wb = load_workbook(input_file, data_only=True)
        for sheet in wb.get_sheet_names():
            sheet_name = sheet
            spreadsheet[sheet_name] = []
            current_sheet = wb.get_sheet_by_name(sheet)
            dict_keys = []
            for c in range(1,current_sheet.max_column + 1):
                dict_keys.append(current_sheet.cell(row=1,column=c).value)
            for r in range (2,current_sheet.max_row + 1):
                temp_dict = {}
                for c in range(1,current_sheet.max_column + 1):
                    temp_dict[dict_keys[c-1]] = current_sheet.cell(row=r,column=c).value
                spreadsheet[sheet_name].append(temp_dict)
    except IOError:
        return (1, "IOError on input file:%s" % input_file)     
    result = spreadsheet
    return result

def listify(f):
    # convert generator functions to functions that return lists
    @contextfilter
    def listify_template(*args, **kwargs):
        return list(f(*args, **kwargs))
    return listify_template


def my_finalize(x):
    if not x:
        warnings.warn('Warning: Empty variable detected!')
    if x=='':
        warnings.warn('Warning: Empty string detected. Possible undefined variable ')
    return x

def render_template(template, config):
    """Render the jinja2 template"""
    env = Environment(loader=FileSystemLoader('./'), trim_blocks=True, lstrip_blocks=True, finalize=my_finalize)
    FILTERS_TO_LISTIFY = [
                        "map",
                        "select",
                        "selectattr",
                        "reject",
                        "rejectattr"
    ]
    for filt in FILTERS_TO_LISTIFY:
        env.filters[filt] = listify(env.filters[filt])
    tests = inspect.getmembers(custom_tests, inspect.isfunction)
    filters = inspect.getmembers(custom_filters, inspect.isfunction)
    for test_name, test_function in tests:
        env.tests[test_name] = test_function
    for filter_name,  filter_function in filters:
        env.filters[filter_name] = filter_function
    template = env.get_template(template)
    config['workbook'] = config
    return template.render(**config)


def main():
    """Main function"""
    args = get_args()
    config = read_xls_dict(args.file)

    print render_template(args.template, config)

if __name__ == '__main__':
    main()
